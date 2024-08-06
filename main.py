import telebot
import openpyxl

TOKEN = '6659393753:AAE8_04pnP7c4L-LqHK-SJ5HmzJxK_1LFLU'
bot = telebot.TeleBot(TOKEN)

# Path to the Excel file
excel_path = '/Users/nasiba/Desktop/Attendance_Bot/partial1/attendance.xlsx'

# Dictionary to store attendance status for each user
attendance = {}

def create_excel():
    # Create a new Excel file if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(['Username', 'Attendance'])
        workbook.save(excel_path)

def update_attendance_excel():
    # Open the Excel file and update attendance data
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    for user, status in attendance.items():
        for row in sheet.iter_rows(min_row=2, max_col=2):
            if row[0].value == user:
                row[1].value = status
                break  # Exit loop once the user is found
    workbook.save(excel_path)


def load_attendance_from_excel():
    # Load attendance data from the Excel file
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        attendance[row[0]] = row[1]

@bot.message_handler(commands=['start'])
def start(message):
    # Extract user's username
    username = message.from_user.username

    # Update user's attendance status to "present"
    attendance[username] = "present"
    bot.reply_to(message, f"Добро пожаловать, {username}! Вас отметили ✅")

    # Update attendance in Excel file
    update_attendance_excel()

if __name__ == "__main__":
    create_excel()  # Create Excel file if it doesn't exist
    load_attendance_from_excel()  # Load attendance data from Excel file
    bot.polling(none_stop=True)
    